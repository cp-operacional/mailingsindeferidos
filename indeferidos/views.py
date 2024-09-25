from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction
from .models import Indeferidos, IndeferidosDone
from .serializers import IndeferidosSerializer
from django.db.models import Q
from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from django.db.models import Count

class IndeferidosListView(APIView):
    def post(self, request):
        campos = request.data.get('campos', [])
        quantidade = request.data.get('quantidade', 10)
        filtros = request.data.get('filtros', [])

        if not campos:
            return Response({"erro": "Nenhum campo especificado"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            cpfs_processados = set(IndeferidosDone.objects.values_list('cpf', flat=True))
            
            resultado = []
            cpfs_para_inserir = []
            
            for filtro in filtros:
                if len(resultado) >= quantidade:
                    break

                queryset = Indeferidos.objects.exclude(cpf__in=cpfs_processados)
                
                limite_filtro = filtro.pop('limite', None)
                
                for campo, valores in filtro.items():
                    if valores:
                        if campo == 'competencia_indeferimento':
                            if len(valores) == 1:
                                queryset = queryset.filter(competencia_indeferimento=valores[0])
                            elif len(valores) == 2:
                                queryset = queryset.filter(competencia_indeferimento__range=(min(valores), max(valores)))
                        elif campo == 'uf_municipio':
                            q_objects = Q()
                            for uf, municipios in valores.items():
                                q_objects |= Q(uf=uf, municipio__in=municipios)
                            queryset = queryset.filter(q_objects)
                        else:
                            queryset = queryset.filter(**{f"{campo}__in": valores})

                limite_atual = min(
                    limite_filtro or quantidade,
                    quantidade - len(resultado)
                )
                
                queryset = queryset[:limite_atual]
                serializer = IndeferidosSerializer(queryset, many=True)

                for item in serializer.data:
                    if len(resultado) >= quantidade:
                        break
                    resultado_item = {campo: item[campo] for campo in campos if campo in item}
                    resultado.append(resultado_item)
                    if 'cpf' in item:
                        cpfs_para_inserir.append(IndeferidosDone(cpf=item['cpf']))

            IndeferidosDone.objects.bulk_create(cpfs_para_inserir, ignore_conflicts=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Indeferidos"

        for col, campo in enumerate(campos, start=1):
            ws.cell(row=1, column=col, value=campo)

        for row, item in enumerate(resultado, start=2):
            for col, campo in enumerate(campos, start=1):
                ws.cell(row=row, column=col, value=item.get(campo, ''))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=indeferidos.xlsx'

        return response

class IndeferidosCountView(APIView):
    def post(self, request):
        filtros = request.data.get('filtros', [])

        queryset = Indeferidos.objects.none()
        ufs_para_contar = set()
        municipios_para_contar = set()

        for filtro in filtros:
            filtro_queryset = Indeferidos.objects.all()
            for campo, valores in filtro.items():
                if valores:
                    if campo == 'competencia_indeferimento':
                        if len(valores) == 1:
                            filtro_queryset = filtro_queryset.filter(competencia_indeferimento=valores[0])
                        elif len(valores) == 2:
                            filtro_queryset = filtro_queryset.filter(competencia_indeferimento__range=(min(valores), max(valores)))
                    elif campo == 'uf_municipio':
                        q_objects = Q()
                        for uf, municipios in valores.items():
                            q_objects |= Q(uf=uf, municipio__in=municipios)
                            municipios_para_contar.update(municipios)
                        filtro_queryset = filtro_queryset.filter(q_objects)
                    elif campo == 'uf':
                        filtro_queryset = filtro_queryset.filter(uf__in=valores)
                        ufs_para_contar.update(valores)
                    else:
                        filtro_queryset = filtro_queryset.filter(**{f"{campo}__in": valores})
            queryset = queryset | filtro_queryset

        total = queryset.count()
        resultado = {
            "count": {
                "total": total
            },
            "query": request.data
        }

        if ufs_para_contar:
            uf_counts = queryset.filter(uf__in=ufs_para_contar).values('uf').annotate(count=Count('uf'))
            for item in uf_counts:
                resultado["count"][item['uf']] = item['count']

        if municipios_para_contar:
            municipio_counts = queryset.filter(municipio__in=municipios_para_contar).values('municipio').annotate(count=Count('cpf'))
            for item in municipio_counts:
                resultado["count"][item['municipio']] = item['count']

        return Response(resultado)